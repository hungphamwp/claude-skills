#!/usr/bin/env python3
"""
Generate an HPB Media quote (báo giá) .xlsx from the master template, given a
JSON spec. See SKILL.md for the field list and defaults.

Usage:
    python3 generate_quote.py spec.json output.xlsx
"""
import json
import sys
import textwrap
from copy import copy
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

TEMPLATE = Path(__file__).parent.parent / "assets" / "template.xlsx"

FIRST_ITEM_ROW = 11  # row of the first (and, in the template, only) item row
ITEM_COLS = ["A", "B", "C", "D", "E", "F"]

DEFAULT_TERMS = {
    "thanh_toan": "Thanh toán: Tạm ứng 50% khi triển khai — 50% còn lại khi bàn giao hoàn chỉnh",
    "bao_hanh": "Bảo hành: 30 ngày kể từ ngày bàn giao (lỗi phát sinh từ phía HPB Media)",
    "hieu_luc": "Báo giá có hiệu lực trong 15 ngày kể từ ngày gửi",
    "phat_sinh": "Yêu cầu phát sinh ngoài phạm vi báo giá sẽ được thống nhất và báo giá bổ sung riêng",
    "ghi_chu": "Tạm ứng 50% — 50% khi bàn giao",
}


def estimate_row_height(chi_tiet: str, wrap_width: int = 48,
                         pt_per_line: float = 13.5, padding: float = 15.0,
                         minimum: float = 90.0) -> float:
    """Calibrated against the original template's malware-cleanup row
    (15 wrapped lines -> 196.5pt height)."""
    total_lines = 0
    for line in chi_tiet.split("\n"):
        if line.strip() == "":
            total_lines += 1
        else:
            wrapped = textwrap.wrap(line, width=wrap_width)
            total_lines += max(1, len(wrapped))
    return max(minimum, total_lines * pt_per_line + padding)


def copy_cell_style(src, dst):
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format


def insert_extra_item_rows(ws, extra: int):
    """Make room for `extra` more item rows below the template's single item
    row (row 11), by inserting blank rows at row 12 and then manually fixing
    up everything openpyxl's insert_rows() does NOT shift on its own:
    merged-cell ranges and floating images (the QR code). Cell values/styles
    for individual (non-merged) cells DO shift correctly on their own.
    """
    if extra <= 0:
        return
    insertion_row = FIRST_ITEM_ROW + 1  # 12

    merges_to_shift = [str(m) for m in list(ws.merged_cells.ranges)
                        if m.min_row >= insertion_row]
    for m in merges_to_shift:
        ws.unmerge_cells(m)

    # image anchors are 0-indexed; insertion_row is 1-indexed
    images_to_shift = [img for img in ws._images
                        if img.anchor._from.row >= insertion_row - 1]

    ws.insert_rows(insertion_row, amount=extra)

    for m in merges_to_shift:
        min_col, min_row, max_col, max_row = range_boundaries(m)
        ws.merge_cells(
            f"{get_column_letter(min_col)}{min_row + extra}:"
            f"{get_column_letter(max_col)}{max_row + extra}"
        )

    for img in images_to_shift:
        img.anchor._from.row += extra
        to = getattr(img.anchor, "to", None)
        if to is not None:
            to.row += extra


def normalize_items(spec: dict) -> list:
    """Accept either `hang_muc_items` (a list, for multi-row quotes) or the
    legacy single-item shorthand (`hang_muc`/`chi_tiet`/`gia`/`ghi_chu`)."""
    items = spec.get("hang_muc_items")
    if items:
        return items
    return [{
        "ten_hang_muc": spec["hang_muc"],
        "chi_tiet": spec["chi_tiet"],
        "gia": spec["gia"],
        "ghi_chu": spec.get("ghi_chu", DEFAULT_TERMS["ghi_chu"]),
    }]


def build(spec: dict, output_path: str):
    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb.active

    today = date.today()
    so_bg = spec.get("so_bg") or f"BG-WEB-{today.strftime('%d%m')}"
    ngay = spec.get("ngay") or today.strftime("%d/%m/%Y")

    ws["A2"] = f"Số BG: {so_bg}   |   Ngày: {ngay}"
    ws["A3"] = spec["tieu_de"]
    ws["A4"] = spec["phu_de"]

    ws["C6"] = spec["khach_hang"]
    ws["C7"] = spec["loai_dich_vu"]
    ws["C8"] = spec["tien_do"]

    items = normalize_items(spec)
    n = len(items)
    insert_extra_item_rows(ws, n - 1)

    for i, item in enumerate(items):
        row = FIRST_ITEM_ROW + i
        if i > 0:
            for col in ITEM_COLS:
                copy_cell_style(ws[f"{col}{FIRST_ITEM_ROW}"], ws[f"{col}{row}"])
        ws[f"A{row}"] = i + 1
        ws[f"B{row}"] = item["ten_hang_muc"]
        ws[f"C{row}"] = item["chi_tiet"]
        ws[f"D{row}"] = item["gia"]
        ws[f"E{row}"] = f"=D{row}"
        ws[f"F{row}"] = item.get("ghi_chu", DEFAULT_TERMS["ghi_chu"])
        ws.row_dimensions[row].height = estimate_row_height(item["chi_tiet"])

    last_item_row = FIRST_ITEM_ROW + n - 1
    total_row = last_item_row + 2  # one spacer row between items and TỔNG CỘNG
    ws[f"E{total_row}"] = f"=SUM(E{FIRST_ITEM_ROW}:E{last_item_row})"

    terms_offset = n - 1
    ws[f"A{16 + terms_offset}"] = spec["thoi_gian_thuc_hien"]
    ws[f"A{17 + terms_offset}"] = spec.get("thanh_toan", DEFAULT_TERMS["thanh_toan"])
    ws[f"A{18 + terms_offset}"] = spec.get("bao_hanh", DEFAULT_TERMS["bao_hanh"])
    ws[f"A{19 + terms_offset}"] = spec.get("hieu_luc", DEFAULT_TERMS["hieu_luc"])
    ws[f"A{20 + terms_offset}"] = spec.get("phat_sinh", DEFAULT_TERMS["phat_sinh"])
    # The contact footer (last row) is left untouched — HPB Media's own info.

    wb.save(output_path)
    print(f"saved -> {output_path} ({n} hạng mục, dòng TỔNG CỘNG = {total_row})")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python3 generate_quote.py spec.json output.xlsx", file=sys.stderr)
        sys.exit(1)
    with open(sys.argv[1], encoding="utf-8") as f:
        spec = json.load(f)
    build(spec, sys.argv[2])
