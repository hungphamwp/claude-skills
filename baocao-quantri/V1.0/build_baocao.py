"""
build_baocao.py — HPB Media Báo Cáo Quản Trị Website
Xuất file xlsx báo cáo tháng chuyên nghiệp gửi khách hàng.
"""

import os
import json
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

SO_LIEN_TUC_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "so-lien-tuc.json")
SKILL_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")

# ── Màu sắc ──────────────────────────────────────────────────────────────────
BLUE       = "0068B8"
ORANGE     = "F97316"
GREEN      = "16A34A"
RED        = "DC2626"
YELLOW     = "CA8A04"
WHITE      = "FFFFFF"
LIGHT_BLUE = "EBF4FB"
LIGHT_GRAY = "F5F5F5"
MID_GRAY   = "D9D9D9"
DARK_TEXT  = "1A1A1A"
SUB_TEXT   = "555555"
GREEN_LIGHT= "DCFCE7"
YELLOW_LIGHT="FEF9C3"
RED_LIGHT  = "FEE2E2"

STATUS_COLOR = {
    "Hoàn thành" : (GREEN,       GREEN_LIGHT),
    "Đang xử lý" : (YELLOW,      YELLOW_LIGHT),
    "Chờ xử lý"  : (SUB_TEXT,    LIGHT_GRAY),
    "Lỗi"        : (RED,         RED_LIGHT),
}

def _font(name="Arial", size=10, bold=False, color=DARK_TEXT, italic=False):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border(top=True, bottom=True, left=True, right=True, color="CCCCCC"):
    s = Side(style="thin", color=color)
    n = Side(style=None)
    return Border(
        top=s if top else n, bottom=s if bottom else n,
        left=s if left else n, right=s if right else n,
    )

def _mc(ws, rng, val=None, fnt=None, aln=None, fll=None, brd=None):
    ws.merge_cells(rng)
    c = ws[rng.split(":")[0]]
    if val is not None: c.value = val
    if fnt: c.font = fnt
    if aln: c.alignment = aln
    if fll: c.fill = fll
    if brd: c.border = brd
    return c

def _set_row(ws, r, height):
    ws.row_dimensions[r].height = height

def next_so_bao_cao(json_path=SO_LIEN_TUC_PATH):
    try:
        with open(json_path, encoding="utf-8") as f:
            d = json.load(f)
        so_cu = d.get("so_bao_cao_cuoi", "BC-QT-0000")
        parts = so_cu.rsplit("-", 1)
        if len(parts) == 2 and parts[1].isdigit():
            pad = max(len(parts[1]), 4)
            new_so = f"{parts[0]}-{str(int(parts[1]) + 1).zfill(pad)}"
        else:
            new_so = so_cu + "-001"
        d["so_bao_cao_cuoi"] = new_so
        d["cap_nhat_lan_cuoi"] = date.today().strftime("%d/%m/%Y")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(d, f, ensure_ascii=False, indent=2)
        return new_so
    except Exception as e:
        print(f"⚠️  {e}")
        return "BC-QT-0001"

def _embed_logo(ws):
    logo_path = os.path.join(SKILL_DIR, "logohpb.png")
    if not os.path.exists(logo_path):
        ws["A1"].value = "HPB MEDIA"
        ws["A1"].font = _font(size=11, bold=True, color=BLUE)
        ws["A1"].alignment = _align("center", "center")
        return
    from PIL import Image as PILImage
    pil_img = PILImage.open(logo_path).convert("RGBA")
    orig_w, orig_h = pil_img.size
    tmp = "/tmp/logo_baocao.png"
    pil_img.save(tmp, "PNG", optimize=False)
    display_h = 52
    display_w = int(orig_w * display_h / orig_h)
    img = XLImage(tmp)
    img.width = display_w
    img.height = display_h

    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.utils.units import pixels_to_EMU
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    col_ab_px = int((6 + 22) * 7.5)
    row_h_px  = 60
    off_x = pixels_to_EMU(max((col_ab_px - display_w) // 2, 4))
    off_y = pixels_to_EMU(max((row_h_px - display_h) // 2, 4))
    marker = AnchorMarker(col=0, colOff=off_x, row=0, rowOff=off_y)
    anchor = OneCellAnchor(_from=marker, ext=None)
    anchor.ext = XDRPositiveSize2D(pixels_to_EMU(display_w), pixels_to_EMU(display_h))
    img.anchor = anchor
    ws.add_image(img)

def build_baocao(data: dict, output_path: str = None):
    """
    data dict:
      THÔNG TIN BÁO CÁO:
        so_bao_cao      : str   — VD: "BC-QT-0001"
        thang_bao_cao   : str   — VD: "04/2026"
        ngay_lap        : str   — để trống → lấy hôm nay

      THÔNG TIN KHÁCH HÀNG:
        ten_khach_hang  : str
        website         : str   — VD: "abc.com.vn"
        nguoi_lien_he   : str
        sdt_khach       : str
        email_khach     : str

      TỔNG QUAN THÁNG:
        tong_quan: [
          {"chi_so": "Uptime", "gia_tri": "99.9%", "ghi_chu": ""},
          {"chi_so": "Số lần backup", "gia_tri": "4 lần", "ghi_chu": "Mỗi tuần 1 lần"},
          ...
        ]

      CÔNG VIỆC ĐÃ THỰC HIỆN:
        cong_viec: [
          {
            "stt"       : 1,
            "loai"      : "Bảo mật",     # Bảo mật / Nội dung / Kỹ thuật / Cập nhật
            "mo_ta"     : "Cập nhật 5 plugin lên phiên bản mới nhất",
            "ngay"      : "05/04/2026",
            "trang_thai": "Hoàn thành",   # Hoàn thành / Đang xử lý / Chờ xử lý / Lỗi
          },
          ...
        ]

      VẤN ĐỀ PHÁT SINH (có thể rỗng):
        van_de: [
          {"mo_ta": "...", "xu_ly": "...", "trang_thai": "Hoàn thành"},
        ]

      KHUYẾN NGHỊ & KẾ HOẠCH:
        khuyen_nghi: ["...", "..."]
        ke_hoach    : ["...", "..."]
    """
    if output_path is None:
        ten_kh = data.get("ten_khach_hang", "KhachHang").replace(" ", "")[:20]
        thang  = data.get("thang_bao_cao", "").replace("/", "-")
        output_path = os.path.join(".", f"BaoCao-{ten_kh}-T{thang}.xlsx")

    ngay_lap = data.get("ngay_lap") or date.today().strftime("%d/%m/%Y")
    so_bc    = data.get("so_bao_cao", "BC-QT-0001")
    thang    = data.get("thang_bao_cao", "")

    wb = Workbook()
    ws = wb.active
    ws.title = "Báo Cáo"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1

    # Cột: A=STT B=Loại C=Mô tả D=Ngày E=Trạng thái
    for col, w in [("A",6),("B",22),("C",42),("D",14),("E",16)]:
        ws.column_dimensions[col].width = w

    r = 1

    # ── ROW 1: Logo + Thông tin công ty ───────────────────────────────────────
    _set_row(ws, r, 60)
    _mc(ws, "A1:B1", fll=_fill(WHITE), aln=_align("center","center"))
    _embed_logo(ws)

    from openpyxl.cell.text import InlineFont
    from openpyxl.cell.rich_text import TextBlock, CellRichText
    ws.merge_cells("C1:E1")
    c1 = ws["C1"]
    c1.fill = _fill(WHITE)
    c1.alignment = _align("left","center", wrap=True)
    bold_f   = InlineFont(rFont='Arial', b=True,  sz=10, color=DARK_TEXT)
    normal_f = InlineFont(rFont='Arial', b=False, sz=10, color=DARK_TEXT)
    link_f   = InlineFont(rFont='Arial', b=False, sz=10, color=BLUE, u="single")
    c1.value = CellRichText(
        TextBlock(bold_f,   "HPB MEDIA — Thiết kế website & Digital Marketing\n"),
        TextBlock(normal_f, "SĐT: "),
        TextBlock(link_f,   "0913 337 280"),
        TextBlock(normal_f, "  |  Email: "),
        TextBlock(link_f,   "hungphamcqb@gmail.com"),
    )
    r += 1

    # ── ROW 2: Số BC + Ngày ───────────────────────────────────────────────────
    _set_row(ws, r, 18)
    _mc(ws, f"A{r}:E{r}",
        val=f"Số BC: {so_bc}   |   Ngày lập: {ngay_lap}",
        fnt=_font(size=9, color=SUB_TEXT, italic=True),
        aln=_align("right","center"),
        fll=_fill(WHITE))
    r += 1

    # ── ROW 3: Tiêu đề ────────────────────────────────────────────────────────
    _set_row(ws, r, 38)
    _mc(ws, f"A{r}:E{r}",
        val=f"BÁO CÁO QUẢN TRỊ WEBSITE — THÁNG {thang}",
        fnt=_font(size=16, bold=True, color=WHITE),
        aln=_align("center","center"),
        fll=_fill(BLUE))
    r += 1

    # ── ROW 4: Tên khách ──────────────────────────────────────────────────────
    _set_row(ws, r, 22)
    _mc(ws, f"A{r}:E{r}",
        val=f"{data.get('ten_khach_hang','')}   |   {data.get('website','')}",
        fnt=_font(size=11, bold=True, color=BLUE),
        aln=_align("center","center"),
        fll=_fill(LIGHT_BLUE))
    r += 1

    # ── Separator ─────────────────────────────────────────────────────────────
    _set_row(ws, r, 4)
    _mc(ws, f"A{r}:E{r}", fll=_fill(ORANGE))
    r += 1

    # ── SECTION: THÔNG TIN KHÁCH HÀNG ────────────────────────────────────────
    _set_row(ws, r, 22)
    _mc(ws, f"A{r}:E{r}",
        val="THÔNG TIN KHÁCH HÀNG",
        fnt=_font(size=10, bold=True, color=WHITE),
        aln=_align("left","center"),
        fll=_fill(BLUE))
    r += 1

    kh_rows = [
        ("Khách hàng:",   data.get("ten_khach_hang", "")),
        ("Website:",      data.get("website", "")),
        ("Liên hệ:",      f"{data.get('nguoi_lien_he','')}   |   SĐT: {data.get('sdt_khach','')}   |   Email: {data.get('email_khach','')}"),
        ("Kỳ báo cáo:",   f"Tháng {thang}"),
    ]
    for label, val in kh_rows:
        _set_row(ws, r, 20)
        _mc(ws, f"A{r}:B{r}", val=label,
            fnt=_font(bold=True), aln=_align("left","center"),
            fll=_fill(LIGHT_GRAY), brd=_border())
        _mc(ws, f"C{r}:E{r}", val=val,
            fnt=_font(), aln=_align("left","center"),
            fll=_fill(WHITE), brd=_border())
        r += 1

    # ── Separator ─────────────────────────────────────────────────────────────
    _set_row(ws, r, 4)
    _mc(ws, f"A{r}:E{r}", fll=_fill(ORANGE))
    r += 1

    # ── SECTION: TỔNG QUAN THÁNG ─────────────────────────────────────────────
    _set_row(ws, r, 22)
    _mc(ws, f"A{r}:E{r}",
        val="TỔNG QUAN THÁNG",
        fnt=_font(size=10, bold=True, color=WHITE),
        aln=_align("left","center"),
        fll=_fill(BLUE))
    r += 1

    # Header tổng quan
    _set_row(ws, r, 22)
    for col_idx, label in enumerate(["STT","CHỈ SỐ","GIÁ TRỊ","","GHI CHÚ"], 1):
        c = ws[f"{get_column_letter(col_idx)}{r}"]
        c.value = label
        c.font = _font(size=10, bold=True, color=WHITE)
        c.alignment = _align("center","center")
        c.fill = _fill("2563EB")
        c.border = _border()
    ws.merge_cells(f"C{r}:D{r}")
    r += 1

    tong_quan = data.get("tong_quan", [])
    for i, item in enumerate(tong_quan):
        _set_row(ws, r, 22)
        bg = WHITE if i % 2 == 0 else LIGHT_GRAY
        ws[f"A{r}"].value = i + 1
        ws[f"A{r}"].font = _font(); ws[f"A{r}"].alignment = _align("center","center")
        ws[f"A{r}"].fill = _fill(bg); ws[f"A{r}"].border = _border()

        ws[f"B{r}"].value = item.get("chi_so","")
        ws[f"B{r}"].font = _font(bold=True); ws[f"B{r}"].alignment = _align("left","center")
        ws[f"B{r}"].fill = _fill(bg); ws[f"B{r}"].border = _border()

        ws.merge_cells(f"C{r}:D{r}")
        ws[f"C{r}"].value = item.get("gia_tri","")
        ws[f"C{r}"].font = _font(bold=True, color=BLUE)
        ws[f"C{r}"].alignment = _align("center","center")
        ws[f"C{r}"].fill = _fill(bg); ws[f"C{r}"].border = _border()

        ws[f"E{r}"].value = item.get("ghi_chu","")
        ws[f"E{r}"].font = _font(size=9, color=SUB_TEXT, italic=True)
        ws[f"E{r}"].alignment = _align("left","center", wrap=True)
        ws[f"E{r}"].fill = _fill(bg); ws[f"E{r}"].border = _border()
        r += 1

    # ── Separator ─────────────────────────────────────────────────────────────
    _set_row(ws, r, 4)
    _mc(ws, f"A{r}:E{r}", fll=_fill(ORANGE))
    r += 1

    # ── SECTION: CÔNG VIỆC ĐÃ THỰC HIỆN ─────────────────────────────────────
    _set_row(ws, r, 22)
    _mc(ws, f"A{r}:E{r}",
        val="CÔNG VIỆC ĐÃ THỰC HIỆN",
        fnt=_font(size=10, bold=True, color=WHITE),
        aln=_align("left","center"),
        fll=_fill(BLUE))
    r += 1

    _set_row(ws, r, 22)
    for col_idx, label in enumerate(["STT","LOẠI","MÔ TẢ CÔNG VIỆC","NGÀY THỰC HIỆN","TRẠNG THÁI"], 1):
        c = ws[f"{get_column_letter(col_idx)}{r}"]
        c.value = label
        c.font = _font(size=10, bold=True, color=WHITE)
        c.alignment = _align("center","center")
        c.fill = _fill("2563EB")
        c.border = _border()
    r += 1

    cong_viec = data.get("cong_viec", [])
    for i, item in enumerate(cong_viec):
        bg = WHITE if i % 2 == 0 else LIGHT_GRAY
        tt = item.get("trang_thai", "Hoàn thành")
        tt_text_color, tt_bg = STATUS_COLOR.get(tt, (DARK_TEXT, LIGHT_GRAY))

        # Tính row height theo mô tả
        mo_ta = item.get("mo_ta", "")
        lines = max(1, -(-len(mo_ta) // 55))
        _set_row(ws, r, max(22, lines * 14 + 6))

        ws[f"A{r}"].value = item.get("stt", i+1)
        ws[f"A{r}"].font = _font(); ws[f"A{r}"].alignment = _align("center","center")
        ws[f"A{r}"].fill = _fill(bg); ws[f"A{r}"].border = _border()

        ws[f"B{r}"].value = item.get("loai","")
        ws[f"B{r}"].font = _font(bold=True, color=BLUE)
        ws[f"B{r}"].alignment = _align("left","center", wrap=True)
        ws[f"B{r}"].fill = _fill(bg); ws[f"B{r}"].border = _border()

        ws[f"C{r}"].value = mo_ta
        ws[f"C{r}"].font = _font(); ws[f"C{r}"].alignment = _align("left","center", wrap=True)
        ws[f"C{r}"].fill = _fill(bg); ws[f"C{r}"].border = _border()

        ws[f"D{r}"].value = item.get("ngay","")
        ws[f"D{r}"].font = _font(); ws[f"D{r}"].alignment = _align("center","center")
        ws[f"D{r}"].fill = _fill(bg); ws[f"D{r}"].border = _border()

        ws[f"E{r}"].value = tt
        ws[f"E{r}"].font = _font(bold=True, color=tt_text_color)
        ws[f"E{r}"].alignment = _align("center","center")
        ws[f"E{r}"].fill = _fill(tt_bg); ws[f"E{r}"].border = _border()
        r += 1

    # ── SECTION: VẤN ĐỀ PHÁT SINH ────────────────────────────────────────────
    van_de = data.get("van_de", [])
    if van_de:
        _set_row(ws, r, 4)
        _mc(ws, f"A{r}:E{r}", fll=_fill(ORANGE))
        r += 1

        _set_row(ws, r, 22)
        _mc(ws, f"A{r}:E{r}",
            val="VẤN ĐỀ PHÁT SINH & XỬ LÝ",
            fnt=_font(size=10, bold=True, color=WHITE),
            aln=_align("left","center"),
            fll=_fill(BLUE))
        r += 1

        _set_row(ws, r, 22)
        for col_idx, label in enumerate(["STT","VẤN ĐỀ","HƯỚNG XỬ LÝ","","TRẠNG THÁI"], 1):
            c = ws[f"{get_column_letter(col_idx)}{r}"]
            c.value = label
            c.font = _font(size=10, bold=True, color=WHITE)
            c.alignment = _align("center","center")
            c.fill = _fill("2563EB"); c.border = _border()
        ws.merge_cells(f"C{r}:D{r}")
        r += 1

        for i, item in enumerate(van_de):
            bg = WHITE if i % 2 == 0 else LIGHT_GRAY
            tt = item.get("trang_thai","Hoàn thành")
            tt_text_color, tt_bg = STATUS_COLOR.get(tt, (DARK_TEXT, LIGHT_GRAY))
            _set_row(ws, r, 30)

            ws[f"A{r}"].value = i + 1
            ws[f"A{r}"].font = _font(); ws[f"A{r}"].alignment = _align("center","center")
            ws[f"A{r}"].fill = _fill(bg); ws[f"A{r}"].border = _border()

            ws[f"B{r}"].value = item.get("mo_ta","")
            ws[f"B{r}"].font = _font(); ws[f"B{r}"].alignment = _align("left","center", wrap=True)
            ws[f"B{r}"].fill = _fill(bg); ws[f"B{r}"].border = _border()

            ws.merge_cells(f"C{r}:D{r}")
            ws[f"C{r}"].value = item.get("xu_ly","")
            ws[f"C{r}"].font = _font(); ws[f"C{r}"].alignment = _align("left","center", wrap=True)
            ws[f"C{r}"].fill = _fill(bg); ws[f"C{r}"].border = _border()

            ws[f"E{r}"].value = tt
            ws[f"E{r}"].font = _font(bold=True, color=tt_text_color)
            ws[f"E{r}"].alignment = _align("center","center")
            ws[f"E{r}"].fill = _fill(tt_bg); ws[f"E{r}"].border = _border()
            r += 1

    # ── Separator ─────────────────────────────────────────────────────────────
    _set_row(ws, r, 4)
    _mc(ws, f"A{r}:E{r}", fll=_fill(ORANGE))
    r += 1

    # ── SECTION: KHUYẾN NGHỊ & KẾ HOẠCH ─────────────────────────────────────
    _set_row(ws, r, 22)
    _mc(ws, f"A{r}:E{r}",
        val="KHUYẾN NGHỊ & KẾ HOẠCH THÁNG TỚI",
        fnt=_font(size=10, bold=True, color=WHITE),
        aln=_align("left","center"),
        fll=_fill(BLUE))
    r += 1

    # Khuyến nghị
    khuyen_nghi = data.get("khuyen_nghi", [])
    if khuyen_nghi:
        _set_row(ws, r, 20)
        _mc(ws, f"A{r}:E{r}",
            val="Khuyến nghị:",
            fnt=_font(bold=True, color=BLUE),
            aln=_align("left","center"),
            fll=_fill(LIGHT_BLUE))
        r += 1
        for i, item in enumerate(khuyen_nghi):
            _set_row(ws, r, 20)
            bg = WHITE if i % 2 == 0 else LIGHT_GRAY
            _mc(ws, f"A{r}:E{r}",
                val=f"  • {item}",
                fnt=_font(),
                aln=_align("left","center", wrap=True),
                fll=_fill(bg),
                brd=_border(top=False, left=False, right=False))
            r += 1

    # Kế hoạch
    ke_hoach = data.get("ke_hoach", [])
    if ke_hoach:
        _set_row(ws, r, 20)
        _mc(ws, f"A{r}:E{r}",
            val="Kế hoạch tháng tới:",
            fnt=_font(bold=True, color=BLUE),
            aln=_align("left","center"),
            fll=_fill(LIGHT_BLUE))
        r += 1
        for i, item in enumerate(ke_hoach):
            _set_row(ws, r, 20)
            bg = WHITE if i % 2 == 0 else LIGHT_GRAY
            _mc(ws, f"A{r}:E{r}",
                val=f"  • {item}",
                fnt=_font(),
                aln=_align("left","center", wrap=True),
                fll=_fill(bg),
                brd=_border(top=False, left=False, right=False))
            r += 1

    # ── Footer ────────────────────────────────────────────────────────────────
    _set_row(ws, r, 4)
    _mc(ws, f"A{r}:E{r}", fll=_fill(MID_GRAY))
    r += 1

    _set_row(ws, r, 22)
    ws.merge_cells(f"A{r}:E{r}")
    fc = ws[f"A{r}"]
    fc.fill = _fill(BLUE)
    fc.alignment = _align("center","center")
    from openpyxl.cell.text import InlineFont
    from openpyxl.cell.rich_text import TextBlock, CellRichText
    fw = InlineFont(rFont="Arial", b=True, sz=10, color=WHITE)
    fc.value = CellRichText(
        TextBlock(fw, "Báo cáo được lập bởi HPB Media  |  Phạm Văn Hưng  |  "),
        TextBlock(fw, "0913 337 280"),
        TextBlock(fw, "  |  hungphamcqb@gmail.com"),
    )
    fc.hyperlink = "tel:+840913337280"

    ws.print_area = f"A1:E{r}"
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    SO_BC = next_so_bao_cao()
    DATA = {
        "so_bao_cao"    : SO_BC,
        "thang_bao_cao" : "04/2026",
        "ngay_lap"      : "",
        "ten_khach_hang": "CÔNG TY TNHH ABC",
        "website"       : "abc.com.vn",
        "nguoi_lien_he" : "Nguyễn Văn A",
        "sdt_khach"     : "0901 234 567",
        "email_khach"   : "a@abc.com.vn",
        "tong_quan": [
            {"chi_so": "Uptime website",          "gia_tri": "99.9%",   "ghi_chu": "Không có sự cố nào trong tháng"},
            {"chi_so": "Số lần backup",            "gia_tri": "4 lần",   "ghi_chu": "Backup tự động mỗi tuần"},
            {"chi_so": "Cập nhật plugin/theme",   "gia_tri": "8 bản",   "ghi_chu": "Đã cập nhật toàn bộ"},
            {"chi_so": "Tình trạng bảo mật",      "gia_tri": "Tốt",     "ghi_chu": "Không phát hiện mã độc"},
            {"chi_so": "Tốc độ tải trang",        "gia_tri": "2.1s",    "ghi_chu": "PageSpeed 85/100"},
        ],
        "cong_viec": [
            {"stt":1, "loai":"Cập nhật",   "mo_ta":"Cập nhật WordPress core lên v6.5.2",                          "ngay":"03/04/2026", "trang_thai":"Hoàn thành"},
            {"stt":2, "loai":"Cập nhật",   "mo_ta":"Cập nhật 8 plugin lên phiên bản mới nhất",                    "ngay":"03/04/2026", "trang_thai":"Hoàn thành"},
            {"stt":3, "loai":"Bảo mật",    "mo_ta":"Quét mã độc toàn bộ website, kết quả: sạch",                  "ngay":"10/04/2026", "trang_thai":"Hoàn thành"},
            {"stt":4, "loai":"Backup",     "mo_ta":"Backup toàn bộ database và file (4 lần trong tháng)",          "ngay":"Hàng tuần",  "trang_thai":"Hoàn thành"},
            {"stt":5, "loai":"Nội dung",   "mo_ta":"Đăng 3 bài viết mới theo yêu cầu khách hàng",                 "ngay":"15/04/2026", "trang_thai":"Hoàn thành"},
            {"stt":6, "loai":"Kỹ thuật",   "mo_ta":"Tối ưu hình ảnh 45 ảnh còn chưa được nén",                    "ngay":"20/04/2026", "trang_thai":"Hoàn thành"},
        ],
        "van_de": [],
        "khuyen_nghi": [
            "Nên nâng cấp gói hosting lên 10GB — dung lượng hiện tại đã dùng 82%",
            "Cân nhắc bổ sung CDN để cải thiện tốc độ tải trang cho người dùng ở xa",
        ],
        "ke_hoach": [
            "Tiếp tục backup tự động hàng tuần",
            "Kiểm tra và cập nhật plugin/theme định kỳ",
            "Đăng nội dung theo lịch biên tập tháng 05/2026",
            "Quét bảo mật định kỳ giữa tháng",
        ],
    }
    out = build_baocao(DATA, output_path=f"/tmp/BaoCao-{SO_BC}.xlsx")
    print(f"✅ Xuất xong: {out}")
